from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

from .models import Element


def insert_cells(ws: Worksheet, ws_row: int, cells: dict):
    """Insert cell value on row of `ws_row`, getted by column letter from `cells`"""
    for column in (ws.iter_cols(min_row=ws_row, min_col=1, max_col=12, max_row=ws_row)):
        for cell in column:
            letter = cell.column_letter
            if not cells.get(cell.column_letter):
                continue
            cell.value = cells[letter]["value"]
            cell.alignment = cells[letter].get("alignment", None)

    ws_row += 1

    return ws, ws_row


def foreman(project):
    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1
    ws3_row = 1

    ws1_index = 1
    ws2_index = 1
    ws3_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Список работ (Бригадир)"
    ws2 = wb.create_sheet("Список материалов (Бригадир)")
    ws3 = wb.create_sheet("Список общий (Бригадир)")

    for stage in stages:
        ws1.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws2.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws3.merge_cells(f"A{ws3_row}:I{ws3_row}")
        ws1[f"A{ws3_row}"] = ws2[f"A{ws3_row}"] = ws3[f"A{ws3_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws2[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws3[f"A{ws3_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1
        ws2_row += 1
        ws3_row += 1

        cells_work = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"}
        }
        cells = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"},
            "F": {"value": "Кол-во"},
            "G": {"value": "доп ед-изм"},
            "H": {"value": "Вес"},
            "I": {"value": "Объем"}
        }

        ws1, ws1_row = insert_cells(ws1, ws1_row, cells_work)
        ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
        ws3, ws3_row = insert_cells(ws3, ws3_row, cells)

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "D": {"value": construction["measure"]}
            }
            ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
            ws3, ws3_row = insert_cells(ws3, ws3_row, cells)

            elements = construction["elements"]
            for element in elements:
                cells_work = {
                        "A": {
                            "value": None,
                            "alignment": Alignment(horizontal="right")
                        },
                        "B": {
                            "value": element["title"]
                        },
                        "C": {
                            "value": element["original_title"]
                        },
                        "D": {
                            "value": element["count"],
                            "alignment": Alignment(horizontal="right")
                        },
                        "E": {
                            "value": element["measure"]
                        }
                    }
                cells = {
                        "A": {
                            "value": None,
                            "alignment": Alignment(horizontal="right")
                        },
                        "B": {
                            "value": element["title"]
                        },
                        "C": {
                            "value": element["original_title"]
                        },
                        "D": {
                            "value": element["count"],
                            "alignment": Alignment(horizontal="right")
                        },
                        "E": {
                            "value": element["measure"]
                        },
                        "F": {
                            "value": element["count"] * element["conversion_rate"],
                            "alignment": Alignment(horizontal="right")
                        },
                        "G": {
                            "value": element["second_measure"]
                        },
                        "H": {
                            "value": element["weight"]
                        },
                        "I": {
                            "value": element["volume"]
                        }
                    }

                if element["type"] == Element.Type.JOB:
                    cells["A"]["value"] = f"{count_construction}.{ws1_index}"
                    ws1, ws1_row = insert_cells(ws1, ws1_row, cells_work)
                    ws1_index += 1
                elif element["type"] == Element.Type.MATERIAL:
                    cells["A"]["value"] = f"{count_construction}.{ws2_index}"
                    ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
                    ws2_index += 1

                cells["A"]["value"] = f"{count_construction}.{ws3_index}"
                ws3, ws3_row = insert_cells(ws3, ws3_row, cells)
                ws3_index +=1 

            count_construction += 1

    return wb


def purchaser(project):
    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1

    ws1_index = 1
    ws2_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Список материалов по этапам (Закупщик)"
    ws2 = wb.create_sheet("Список материалов общий (Закупщик)")

    for stage_index, stage in enumerate(stages):
        ws1.merge_cells(f"A{ws1_row}:F{ws1_row}")
        ws1[f"A{ws1_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws1_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1

        cells = {
            "B": {"value": "Наименование"},
            "C": {"value": "Описание"},
            "D": {"value": "Кол-во"},
            "E": {"value": "ед-изм"},
            "F": {"value": "Кол-во"},
            "G": {"value": "доп ед-изм"},
            "H": {"value": "Цена/ед"},
            "I": {"value": "Сумма"},
            "J": {"value": "Вес"},
            "K": {"value": "Объем"},
            "L": {"value": "Размер"}
        }
        ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
        if stage_index == 0:
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "D": {"value": construction["measure"]}
            }
            ws1, ws1_row = insert_cells(ws1, ws1_row, cells)

            elements = construction["elements"]
            for element in elements:
                cells = {
                    "A": {
                        "value": None,
                        "alignment": Alignment(horizontal="right")
                    },
                    "B": {
                        "value": element["title"]
                    },
                    "C": {
                        "value": element["original_title"]
                    },
                    "D": {
                        "value": element["count"],
                        "alignment": Alignment(horizontal="right")
                    },
                    "E": {
                        "value": element["measure"]
                    },
                    "F": {
                        "value": element["count"] * element["conversion_rate"],
                    },
                    "G": {
                        "value": element["second_measure"]
                    },
                    "H": {
                        "value": element["cost"]
                    },
                    "I": {
                        "value": None
                    },
                    "J": {
                        "value": element["weight"]
                    },
                    "K": {
                        "value": element["volume"]
                    },
                    "L": {
                        "value": element["dimension"]
                    }
                }
                if element["type"] == Element.Type.MATERIAL:
                    cells["A"]["value"] = f"{count_construction}.{ws1_index}"
                    cells["I"]["value"] = f"=D{ws1_row}*H{ws1_row}"
                    ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
                    ws1_index += 1

                cells["A"]["value"] = f"{count_construction}.{ws2_index}"
                cells["I"]["value"] = f"=D{ws2_row}*H{ws2_row}"
                ws2, ws2_row = insert_cells(ws2, ws2_row, cells)
                ws2_index += 1

            count_construction += 1

    return wb


def estimate(project):
    def sum_total_price(ws, ws_row, ws_price_cells, word: str):
        ws.merge_cells(f"A{ws_row}:F{ws_row}")
        ws[f"A{ws_row}"] = word
        ws[f"G{ws_row}"] = f"=SUM({';'.join(ws_price_cells)})"

        return ws, ws_row

    stages = project["stages"]
    ws1_row = 1
    ws2_row = 1

    ws1_index = 1

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Смета"
    ws2 = wb.create_sheet("Смета(сокр)")

    ws1_stage_price_cells = []
    ws2_stage_price_cells = []

    for stage in stages:
        ws1.merge_cells(f"A{ws1_row}:H{ws1_row}")
        ws2.merge_cells(f"A{ws2_row}:H{ws2_row}")
        ws1[f"A{ws1_row}"] = ws2[f"A{ws2_row}"] = f"Этап {stage['order']}. {stage['title']}"
        ws1[f"A{ws1_row}"].alignment = Alignment(horizontal="center")
        ws2[f"A{ws2_row}"].alignment = Alignment(horizontal="center")
        ws1_row += 1
        ws2_row += 1

        cells = {
            "B": {"value": "Наименование"}, 
            "C": {"value": "Кол-во"},
            "D": {"value": "ед-изм"},
            "E": {"value": "Цена/ед"},
            "F": {"value": "Сумма"},
            "G": {"value": "Итого"}
        }
        ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
        ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

        ws1_construction_price_cells = []
        ws2_construction_price_cells = []

        constructions = stage["constructions"]
        for count_construction, construction in enumerate(constructions, start=1):
            cells = {
                "A": {"value": f"{count_construction}. Конструкция"},
                "B": {"value": construction["title"]},
                "C": {"value": construction["count"]},
                "D": {"value": construction["measure"]},
                "F": {"value": f"=SUM(F{ws1_row + 1}:F{ws1_row + len(construction['elements'])})"}
            }
            ws1, ws1_row = insert_cells(ws1, ws1_row, cells)
            cells["F"]["value"] = f"=SUM(F{ws2_row + 1}:F{ws2_row + len(construction['elements'])})"
            ws2, ws2_row = insert_cells(ws2, ws2_row, cells)

            ws1_construction_price_cells.append(f"F{ws1_row-1}")
            ws2_construction_price_cells.append(f"F{ws2_row-1}")

            elements_price = 0

            elements = construction["elements"]
            for element in elements:
                cells = {
                    "A": {
                        "value": f"{count_construction}.{ws1_index}",
                        "alignment": Alignment(horizontal="right")
                    },
                    "B": {
                        "value": element["title"]
                    },
                    "C": {
                        "value": element["count"],
                    },
                    "D": {
                        "value": element["measure"]
                    },
                    "E": {
                        "value": element["cost"]
                    },
                    "F": {
                        "value": f"=E{ws1_row}*C{ws1_row}"
                    }
                }
                ws1, ws1_row = insert_cells(ws1, ws1_row, cells)

                ws1_index += 1

                elements_price += element['cost'] * element['count']

            ws2[f"F{ws2_row - 1}"] = elements_price
            count_construction += 1

        ws1, ws1_row = sum_total_price(ws1, ws1_row, ws1_construction_price_cells, "Итого")
        ws1_stage_price_cells.append(f"G{ws1_row}")
        ws1_row += 1

        ws2, ws2_row = sum_total_price(ws2, ws2_row, ws2_construction_price_cells, "Итого")
        ws2_stage_price_cells.append(f"G{ws2_row}")
        ws2_row += 1

    ws1, ws1_row = sum_total_price(ws1, ws1_row, ws1_stage_price_cells, "Всего")
    ws2, ws2_row = sum_total_price(ws2, ws2_row, ws2_stage_price_cells, "Всего")

    return wb
